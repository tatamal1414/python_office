from re import X
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

ws1 = wb.create_sheet("Worksheet")

for i in range(0,10):
    ws1.append([x+i*10 for x in range(1,11) ])

# Save the file
wb.save("sample.xlsx")